#!/usr/bin/env python3
"""
Government bond yield curve data pipeline.

Pulls daily constant-maturity / spot yield curves from official sources:
  US  - US Treasury, Daily Treasury Par Yield Curve Rates (XML feed)
  UK  - Bank of England Government Liability Curve (nominal spot, daily archive)
        spliced forward with BoE IADB daily par yields (IUDSNPY/IUDMNPY/IUDLNPY)
  DE  - Deutsche Bundesbank BBSIS term structure (Svensson), daily
  JP  - Japan MOF, JGB daily yields for all maturities

Emits rates.json: compact per-market date + tenor arrays, ready to embed in HTML.

Adding a market: append an entry to MARKETS and write a fetch_<code>() returning
{ 'YYYY-MM-DD': {tenor: yield_pct, ...}, ... }
"""

import csv
import io
import json
import os
import re
import sys
import time
import zipfile
import datetime as dt
from urllib.error import HTTPError, URLError
from urllib.request import urlopen, Request

HERE = os.path.dirname(os.path.abspath(__file__))
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")
TIMEOUT = 180
HISTORY_YEARS = 6          # keep enough for the 5y lookback column
# UK curve basis: "spot" (GLC files, works anywhere) or "par" (BoE IADB, blocked
# from datacentre IPs). Cloud runs use spot.
UK_BASIS = os.environ.get("UK_BASIS", "spot").lower()
TODAY = dt.date.today()
START = TODAY - dt.timedelta(days=int(365.25 * HISTORY_YEARS))


RETRY_ATTEMPTS = 3
RETRY_BACKOFF = 1.5        # seconds before the first retry, doubling after


def _fetch(url, headers, timeout):
    """One GET, retried through a transient upstream failure.

    Publishers have moments. On 1 Sep 2026 the Bundesbank answered 400 to all
    six tenor requests inside six seconds and was serving normally an hour
    later; with a single-shot fetch that cost the whole build, because a market
    that drops out entirely is fatal by design. A second attempt is far cheaper
    than a lost window.

    A 404 is not a moment, so it is not retried — that would just add seconds
    to every run against a URL that has genuinely moved. A real outage still
    exhausts the attempts and fails as loudly as before.
    """
    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            with urlopen(Request(url, headers=headers), timeout=timeout) as r:
                return r.status, (r.headers.get("Content-Type") or "?"), r.read()
        except HTTPError as e:
            if e.code == 404 or attempt == RETRY_ATTEMPTS:
                raise
            why = f"HTTP {e.code}"
        except (URLError, OSError) as e:
            if attempt == RETRY_ATTEMPTS:
                raise
            why = str(e)
        wait = RETRY_BACKOFF * 2 ** (attempt - 1)
        log(f"retrying in {wait:.1f}s after {why} "
            f"(attempt {attempt} of {RETRY_ATTEMPTS}): {url[:90]}")
        time.sleep(wait)


def get(url, binary=False, timeout=TIMEOUT, accept="*/*"):
    headers = {"User-Agent": UA, "Accept": accept}
    # The BoE download links check the referer; sending that referer to every
    # other publisher is at best pointless and at worst reads as a bot.
    if "bankofengland.co.uk" in url:
        headers["Referer"] = "https://www.bankofengland.co.uk/"
    _, _, raw = _fetch(url, headers, timeout)
    return raw if binary else raw.decode("utf-8", "replace")


def get_json(url, timeout=TIMEOUT):
    """Fetch JSON, and say what actually arrived when it is not JSON.

    A publisher that starts serving an interstitial or a block page answers 200
    with HTML, so json.loads reports "Expecting value: line 1 column 1" and
    nothing about the cause. Carry the status, type and opening bytes into the
    error instead — that is the difference between a diagnosable failure and a
    guess.
    """
    status, ctype, body = _fetch(
        url, {"User-Agent": UA, "Accept": "application/json"}, timeout)
    raw = body.decode("utf-8", "replace")
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        head = " ".join(raw[:200].split()) or "<empty body>"
        raise RuntimeError(
            f"{status} {ctype}, {len(raw)} bytes, not JSON ({e}); starts: {head}"
        ) from None


def log(msg):
    print(f"[{dt.datetime.now():%H:%M:%S}] {msg}", file=sys.stderr)


# ---------------------------------------------------------------- US --------
US_TENOR_TAGS = {
    "3M": "BC_3MONTH", "1Y": "BC_1YEAR", "2Y": "BC_2YEAR", "5Y": "BC_5YEAR",
    "10Y": "BC_10YEAR", "20Y": "BC_20YEAR", "30Y": "BC_30YEAR",
}


def fetch_us():
    out = {}
    for year in range(START.year, TODAY.year + 1):
        url = ("https://home.treasury.gov/resource-center/data-chart-center/"
               "interest-rates/pages/xml?data=daily_treasury_yield_curve"
               f"&field_tdr_date_value={year}")

        try:
            xml = get(url)
        except Exception as e:                                    # noqa: BLE001
            log(f"US {year} failed: {e}")
            continue
        for entry in re.findall(r"<m:properties>(.*?)</m:properties>", xml, re.S):
            m = re.search(r"<d:NEW_DATE[^>]*>([0-9-]{10})", entry)
            if not m:
                continue
            date = m.group(1)
            row = {}
            for tenor, tag in US_TENOR_TAGS.items():
                v = re.search(rf"<d:{tag}[^>]*>([-0-9.]+)<", entry)
                if v:
                    row[tenor] = float(v.group(1))
            if row:
                out[date] = row
        log(f"US {year}: {len(out)} cumulative obs")
    return out


# ---------------------------------------------------------------- UK --------
UK_GLC_MATS = {"2Y": 2.0, "5Y": 5.0, "10Y": 10.0, "20Y": 20.0, "30Y": 30.0}
# BoE IADB daily par yields used to extend the GLC archive to today
UK_IADB = {"5Y": "IUDSNPY", "10Y": "IUDMNPY", "20Y": "IUDLNPY"}
# tenors with no daily official series -> proxied off the nearest official point
UK_PROXY = {"2Y": "5Y", "30Y": "20Y"}


def _glc_zip():
    cached = os.path.join(HERE, "glcnominalddata.zip")
    if os.path.exists(cached) and os.path.getsize(cached) > 1_000_000:
        age = TODAY - dt.date.fromtimestamp(os.path.getmtime(cached))
        if age.days < 7:
            log("UK: using cached GLC archive")
            return open(cached, "rb").read()
    log("UK: downloading BoE GLC nominal daily archive (~39MB)")
    raw = get("https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
              "yield-curves/glcnominalddata.zip", binary=True)
    with open(cached, "wb") as f:
        f.write(raw)
    return raw


def _parse_glc_sheet(ws):
    """Return {date: {tenor: yield}} from a GLC 'spot curve' worksheet."""
    import openpyxl  # noqa: F401  (imported for side-effect clarity)
    col_for = {}
    rows = ws.iter_rows(values_only=True)
    data = {}
    for r in rows:
        if not r:
            continue
        head = r[0]
        if isinstance(head, str) and head.strip().lower().startswith("years"):
            for i, v in enumerate(r):
                if isinstance(v, (int, float)):
                    for tenor, mat in UK_GLC_MATS.items():
                        if abs(float(v) - mat) < 1e-9:
                            col_for[tenor] = i
            continue
        if isinstance(head, dt.datetime):
            d = head.date()
            if d < START:
                continue
            row = {}
            for tenor, i in col_for.items():
                v = r[i] if i < len(r) else None
                if isinstance(v, (int, float)):
                    row[tenor] = round(float(v), 4)
            if row:
                data[d.isoformat()] = row
    return data


def _glc_current_zip():
    """BoE 'latest yield curve data' — static media file, refreshed daily (T+1).

    Unlike the IADB application endpoint this is a plain file download, so it
    works from datacentre IPs where the IADB WAF returns 'Access denied'.
    """
    log("UK: downloading BoE latest-yield-curve-data.zip")
    return get("https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
               "yield-curves/latest-yield-curve-data.zip", binary=True)


def _glc_from_zip(raw, match):
    import openpyxl
    zf = zipfile.ZipFile(io.BytesIO(raw))
    names = [n for n in zf.namelist() if match(n) and n.endswith(".xlsx")]
    out = {}
    for n in sorted(names):
        log(f"UK: parsing {n}")
        with zf.open(n) as fh:
            wb = openpyxl.load_workbook(io.BytesIO(fh.read()),
                                        read_only=True, data_only=True)
            sheet = next((s for s in wb.sheetnames
                          if s.lower().startswith("4.")
                          or "spot curve" in s.lower()), None)
            if sheet:
                out.update(_parse_glc_sheet(wb[sheet]))
            wb.close()
    return out


def fetch_uk_glc():
    import openpyxl
    raw = _glc_zip()
    zf = zipfile.ZipFile(io.BytesIO(raw))
    names = [n for n in zf.namelist()
             if "Nominal daily" in n and n.endswith(".xlsx")]
    # only the files that can contain dates inside our window
    wanted = []
    for n in names:
        yrs = [int(y) for y in re.findall(r"(\d{4})", n)]
        end = max(yrs) if yrs else 9999
        if "present" in n.lower() or end >= START.year:
            wanted.append(n)
    out = {}
    for n in sorted(wanted):
        log(f"UK: parsing {n}")
        with zf.open(n) as fh:
            wb = openpyxl.load_workbook(io.BytesIO(fh.read()),
                                        read_only=True, data_only=True)
            sheet = next((s for s in wb.sheetnames
                          if s.lower().startswith("4.")
                          or "spot curve" in s.lower()), None)
            if sheet:
                out.update(_parse_glc_sheet(wb[sheet]))
            wb.close()
    try:
        cur = _glc_from_zip(_glc_current_zip(),
                            lambda n: "Nominal daily" in n and "current" in n.lower())
        out.update(cur)
    except Exception as e:                                         # noqa: BLE001
        log(f"UK: current-month GLC unavailable ({e}); archive only")
    log(f"UK GLC: {len(out)} obs, latest {max(out) if out else 'n/a'}")
    return out


def fetch_uk_iadb():
    codes = ",".join(UK_IADB.values())
    url = ("https://www.bankofengland.co.uk/boeapps/database/"
           "_iadb-fromshowcolumns.asp?csv.x=yes"
           f"&Datefrom={START:%d/%b/%Y}&Dateto=now&SeriesCodes={codes}"
           "&CSVF=TN&UsingCodes=Y&VPD=Y&VFD=N")
    txt = get(url)
    rd = csv.DictReader(io.StringIO(txt.replace("\r", "")))
    inv = {v: k for k, v in UK_IADB.items()}
    out = {}
    for row in rd:
        try:
            d = dt.datetime.strptime(row["DATE"].strip(), "%d %b %Y").date()
        except Exception:                                          # noqa: BLE001
            continue
        vals = {}
        for code, tenor in inv.items():
            v = (row.get(code) or "").strip()
            if v:
                try:
                    vals[tenor] = float(v)
                except ValueError:
                    pass
        if vals:
            out[d.isoformat()] = vals
    log(f"UK IADB: {len(out)} obs, latest {max(out) if out else 'n/a'}")
    return out


def fetch_uk(manual=None):
    """Consistent par-basis UK curve on the BoE IADB daily grid.

    5Y / 10Y / 20Y are the official BoE daily par yields. 2Y and 30Y have no
    official daily series, so they are derived as the nearest official par
    yield plus the officially-published GLC spot spread for that gap:

        2Y  = IUDSNPY + (GLC 2y  - GLC 5y)
        30Y = IUDLNPY + (GLC 30y - GLC 20y)

    Spreads come from the BoE Government Liability Curve archive (daily, but
    refreshed monthly); past its cut-off the last observed spread is carried
    forward, which is stable to ~1-2bp over a few weeks. Series levels are NOT
    adjusted to screen quotes - any basis versus a market redemption yield is
    reported in `basis_bp` so it stays visible rather than being papered over.
    """
    glc = fetch_uk_glc()

    if UK_BASIS == "spot":
        # Declared basis for unattended/cloud runs. The BoE IADB application
        # endpoint that serves par yields blocks datacentre IPs, so the curve is
        # sourced entirely from the GLC nominal spot archive + current-month file
        # (both plain file downloads). Spot, not par: long-end levels sit a few
        # bp off a par redemption yield, and that basis is stated on the page.
        if not glc:
            return {}, {}
        log(f"UK: spot basis (GLC), {len(glc)} obs")
        return glc, {
            "basis_override": "nominal spot (GLC)",
            "source_override": "Bank of England — Government Liability Curve, "
                               "nominal spot (daily archive + current month)",
            "official_daily": sorted(UK_GLC_MATS, key=lambda t: UK_GLC_MATS[t]),
            "derived": {},
            "note": "Sourced from the BoE Government Liability Curve (nominal "
                    "spot). Not directly comparable with the US par curve at the "
                    "long end; the gap is a few bp.",
            "spread_source_through": max(glc),
            "quotes": {},
        }

    try:
        iadb = fetch_uk_iadb()
    except Exception as e:                                         # noqa: BLE001
        log(f"UK IADB failed: {e}")
        iadb = {}
    if not iadb:
        if not glc:
            return {}, {}
        log("UK: IADB unavailable -> GLC spot fallback")
        return glc, {
            "basis_override": "nominal spot (GLC)",
            "official_daily": sorted(UK_GLC_MATS, key=lambda t: UK_GLC_MATS[t]),
            "derived": {},
            "degraded": "IADB par yields unavailable; curve shown on GLC spot "
                        "basis instead.",
            "spread_source_through": max(glc),
            "quotes": {},
        }

    spread_dates = sorted(glc)
    last_spreads = None
    spreads = {}
    for d in spread_dates:
        g = glc[d]
        if all(k in g for k in ("2Y", "5Y", "20Y", "30Y")):
            last_spreads = {"2Y": g["2Y"] - g["5Y"], "30Y": g["30Y"] - g["20Y"]}
            spreads[d] = last_spreads

    out = {}
    for d in sorted(iadb):
        row = {t: iadb[d][t] for t in UK_IADB if t in iadb[d]}
        sp = spreads.get(d) or last_spreads
        if sp:
            if "5Y" in row:
                row["2Y"] = round(row["5Y"] + sp["2Y"], 4)
            if "20Y" in row:
                row["30Y"] = round(row["20Y"] + sp["30Y"], 4)
        if row:
            out[d] = row

    meta = {
        "official_daily": ["5Y", "10Y", "20Y"],
        "derived": {"2Y": "5Y par + GLC 2s5s spot spread",
                    "30Y": "20Y par + GLC 20s30s spot spread"},
        "spread_source_through": max(spreads) if spreads else None,
        "quotes": {},
    }
    latest = max(out)
    for tenor, q in ((manual or {}).get("UK") or {}).items():
        try:
            level = float(q["value"])
        except (KeyError, TypeError, ValueError):
            continue
        ref = out.get(q.get("date"), out[latest]).get(tenor)
        meta["quotes"][tenor] = {
            "value": level, "date": q.get("date"),
            "source": q.get("source", "market quote"),
            "basis_bp": round((level - ref) * 100, 1) if ref is not None else None,
        }
    return out, meta


# ---------------------------------------------------------------- DE --------
DE_SERIES = {"1Y": "R01XX", "2Y": "R02XX", "5Y": "R05XX", "10Y": "R10XX",
             "20Y": "R20XX", "30Y": "R30XX"}


def fetch_de():
    out = {}
    for tenor, code in DE_SERIES.items():
        url = ("https://api.statistiken.bundesbank.de/rest/download/BBSIS/"
               f"D.I.ZST.ZI.EUR.S1311.B.A604.{code}.R.A.A._Z._Z.A"
               "?format=csv&lang=en")
        try:
            txt = get(url)
        except Exception as e:                                     # noqa: BLE001
            log(f"DE {tenor} failed: {e}")
            continue
        n = 0
        for line in txt.replace("\r", "").split("\n"):
            parts = line.split(",")
            if len(parts) < 2 or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", parts[0]):
                continue
            if parts[0] < START.isoformat():
                continue
            try:
                v = float(parts[1])
            except ValueError:
                continue
            out.setdefault(parts[0], {})[tenor] = v
            n += 1
        log(f"DE {tenor}: {n} obs")
    return out


# ---------------------------------------------------------------- JP --------
JP_TENORS = ["1Y", "2Y", "5Y", "10Y", "20Y", "30Y", "40Y"]


def fetch_jp():
    # historical/jgbcme_all.csv ends at the previous month-end; jgbcme.csv
    # carries the current month, so both are needed.
    txt = ""
    for u in ("historical/jgbcme_all.csv", "jgbcme.csv"):
        try:
            txt += get("https://www.mof.go.jp/english/policy/jgbs/reference/"
                       f"interest_rate/{u}") + "\n"
        except Exception as e:                                     # noqa: BLE001
            log(f"JP {u} failed: {e}")
    return _parse_jp(txt)


def _parse_jp(txt):
    lines = [l for l in txt.replace("\r", "").split("\n") if l.strip()]
    hdr_i = next(i for i, l in enumerate(lines) if l.startswith("Date,"))
    cols = lines[hdr_i].split(",")
    idx = {t: cols.index(t) for t in JP_TENORS if t in cols}
    out = {}
    for line in lines[hdr_i + 1:]:
        p = line.split(",")
        try:
            d = dt.datetime.strptime(p[0].strip(), "%Y/%m/%d").date()
        except Exception:                                          # noqa: BLE001
            continue
        if d < START:
            continue
        row = {}
        for t, i in idx.items():
            if i < len(p):
                v = p[i].strip()
                if v and v != "-":
                    try:
                        row[t] = float(v)
                    except ValueError:
                        pass
        if row:
            out[d.isoformat()] = row
    log(f"JP: {len(out)} obs, latest {max(out) if out else 'n/a'}")
    return out


# ------------------------------------------------- context: gold & reals ----
US_REAL_TAGS = {"5Y": "TC_5YEAR", "10Y": "TC_10YEAR",
                "20Y": "TC_20YEAR", "30Y": "TC_30YEAR"}


def fetch_us_real():
    """US Treasury daily REAL (TIPS) par yield curve — official, same publisher
    as the nominal curve, so nominal minus real is a clean breakeven."""
    out = {}
    for year in range(START.year, TODAY.year + 1):
        url = ("https://home.treasury.gov/resource-center/data-chart-center/"
               "interest-rates/pages/xml?data=daily_treasury_real_yield_curve"
               f"&field_tdr_date_value={year}")
        try:
            xml = get(url)
        except Exception as e:                                     # noqa: BLE001
            log(f"US real {year} failed: {e}")
            continue
        for entry in re.findall(r"<m:properties>(.*?)</m:properties>", xml, re.S):
            m = re.search(r"<d:NEW_DATE[^>]*>([0-9-]{10})", entry)
            if not m:
                continue
            row = {}
            for tenor, tag in US_REAL_TAGS.items():
                v = re.search(rf"<d:{tag}[^>]*>([-0-9.]+)<", entry)
                if v:
                    row[tenor] = float(v.group(1))
            if row:
                out[m.group(1)] = row
    log(f"US real: {len(out)} obs, latest {max(out) if out else 'n/a'}")
    return out


def fetch_gold():
    """LBMA Gold Price PM auction — the settlement benchmark, USD/GBP/EUR."""
    raw = get_json("https://prices.lbma.org.uk/json/gold_pm.json")
    out = {}
    for r in raw:
        d = r.get("d")
        if not d or d < START.isoformat():
            continue
        v = r.get("v") or []
        row = {}
        for i, ccy in enumerate(("USD", "GBP", "EUR")):
            if i < len(v) and isinstance(v[i], (int, float)):
                row[ccy] = float(v[i])
        if row.get("USD"):
            out[d] = row
    log(f"Gold: {len(out)} obs, latest {max(out) if out else 'n/a'}")
    return out


def build_context(us_nominal):
    """Returns (ctx, errors). A context series that drops out is not fatal —
    the curves are the point of the page — but it must not vanish quietly, so
    every failure is carried out to the payload and shown on the page."""
    ctx, errors = {}, []
    try:
        real = fetch_us_real()
    except Exception as e:                                         # noqa: BLE001
        log(f"US real FAILED: {e}")
        errors.append({"series": "US real yields (TIPS)", "detail": str(e)})
        real = {}
    if real:
        dates = sorted(real)
        tenors = [t for t in US_REAL_TAGS if any(t in real[d] for d in dates)]
        ctx["usreal"] = {
            "label": "US real yields (TIPS)", "unit": "%",
            "dates": dates, "tenors": tenors,
            "series": {t: [real[d].get(t) for d in dates] for t in tenors},
            "asof": dates[-1],
            "source": "US Treasury — Daily Real Yield Curve",
            "source_url": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_real_yield_curve",
        }
        # breakeven inflation = nominal par - real par, same publisher, same date
        be_dates, be = [], {t: [] for t in tenors}
        for d in dates:
            if d not in us_nominal:
                continue
            row = {}
            for t in tenors:
                n, r = us_nominal[d].get(t), real[d].get(t)
                row[t] = round(n - r, 4) if (n is not None and r is not None) else None
            if any(v is not None for v in row.values()):
                be_dates.append(d)
                for t in tenors:
                    be[t].append(row[t])
        if be_dates:
            ctx["breakeven"] = {
                "label": "US breakeven inflation", "unit": "%",
                "dates": be_dates, "tenors": tenors, "series": be,
                "asof": be_dates[-1],
                "source": "US Treasury — nominal par less real par",
                "source_url": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_real_yield_curve",
            }
    if not real:
        errors.append({"series": "US breakeven inflation",
                       "detail": "derived from the real curve, which failed"})
    try:
        gold = fetch_gold()
    except Exception as e:                                         # noqa: BLE001
        log(f"Gold FAILED: {e}")
        errors.append({"series": "Gold, LBMA PM auction", "detail": str(e)})
        gold = {}
    if gold:
        dates = sorted(gold)
        ccys = [c for c in ("USD", "GBP", "EUR") if any(c in gold[d] for d in dates)]
        ctx["gold"] = {
            "label": "Gold, LBMA PM auction", "unit": "ccy",
            "dates": dates, "tenors": ccys,
            "series": {c: [gold[d].get(c) for d in dates] for c in ccys},
            "asof": dates[-1],
            "source": "LBMA — Gold Price PM auction",
            "source_url": "https://www.lbma.org.uk/prices-and-data/precious-metal-prices",
        }
    return ctx, errors


# ------------------------------------------------------------- assemble -----
MARKETS = [
    {"code": "US", "name": "United States", "flag": "US", "ccy": "USD",
     "tenors": ["3M", "1Y", "2Y", "5Y", "10Y", "20Y", "30Y"],
     "key": ["2Y", "30Y"],
     "spreads": [["2Y", "10Y"], ["2Y", "30Y"], ["5Y", "30Y"], ["10Y", "30Y"]],
     "source": "US Treasury — Daily Par Yield Curve",
     "source_url": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve",
     "basis": "par yield, constant maturity"},
    {"code": "UK", "name": "United Kingdom", "flag": "GB", "ccy": "GBP",
     "tenors": ["2Y", "5Y", "10Y", "20Y", "30Y"],
     "key": ["5Y", "30Y"],
     "spreads": [["2Y", "10Y"], ["5Y", "30Y"], ["10Y", "30Y"]],
     "source": "Bank of England GLC nominal spot curve + IADB daily par yields",
     "source_url": "https://www.bankofengland.co.uk/statistics/yield-curves",
     "basis": "nominal spot (Svensson-style GLC)"},
    {"code": "DE", "name": "Germany", "flag": "DE", "ccy": "EUR",
     "tenors": ["1Y", "2Y", "5Y", "10Y", "20Y", "30Y"],
     "key": ["2Y", "30Y"],
     "spreads": [["2Y", "10Y"], ["5Y", "30Y"], ["10Y", "30Y"]],
     "source": "Deutsche Bundesbank — BBSIS term structure (Svensson)",
     "source_url": "https://www.bundesbank.de/en/statistics/money-and-capital-markets/interest-rates-and-yields",
     "basis": "zero-coupon spot, listed Federal securities"},
    {"code": "JP", "name": "Japan", "flag": "JP", "ccy": "JPY",
     "tenors": ["1Y", "2Y", "5Y", "10Y", "20Y", "30Y", "40Y"],
     "key": ["2Y", "30Y"],
     "spreads": [["2Y", "10Y"], ["5Y", "30Y"], ["10Y", "30Y"], ["30Y", "40Y"]],
     "source": "Japan Ministry of Finance — JGB daily interest rates",
     "source_url": "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/",
     "basis": "compound yield, benchmark maturities"},
]

FETCHERS = {"US": fetch_us, "DE": fetch_de, "JP": fetch_jp}


def main():
    manual = {}
    mq = os.path.join(HERE, "manual_quotes.json")
    if os.path.exists(mq):
        manual = json.load(open(mq))
    policy = json.load(open(os.path.join(HERE, "policy_rates.json")))

    out_markets = []
    us_nominal = {}
    for spec in MARKETS:
        code = spec["code"]
        log(f"=== {code} ===")
        try:
            if code == "UK":
                data, meta = fetch_uk(manual)
            else:
                data, meta = FETCHERS[code](), {}
            if code == "US":
                us_nominal = data
        except Exception as e:                                     # noqa: BLE001
            log(f"{code} FAILED: {e}")
            continue
        if not data:
            log(f"{code}: no data")
            continue
        dates = sorted(data)
        tenors = [t for t in spec["tenors"] if any(t in data[d] for d in dates)]
        series = {t: [data[d].get(t) for d in dates] for t in tenors}
        m = {k: v for k, v in spec.items()}
        m["basis"] = meta.get("basis_override", m["basis"])
        m["source"] = meta.get("source_override", m["source"])
        m.update({"dates": dates, "tenors": tenors, "series": series,
                  "asof": dates[-1], "meta": meta,
                  "policy": policy.get(code, {})})
        out_markets.append(m)
        log(f"{code}: {len(dates)} dates, tenors {tenors}, asof {dates[-1]}")

    log("=== context (US real yields, breakevens, gold) ===")
    context, context_errors = build_context(us_nominal)

    commentary = {}
    cp = os.path.join(HERE, "commentary.json")
    if os.path.exists(cp):
        try:
            commentary = json.load(open(cp))
        except Exception as e:                                     # noqa: BLE001
            log(f"commentary.json unreadable: {e}")

    payload = {
        "generated": dt.datetime.now(dt.timezone.utc)
                       .strftime("%Y-%m-%dT%H:%M:%SZ"),
        "history_years": HISTORY_YEARS,
        "markets": out_markets,
        "context": context,
        "context_errors": context_errors,
        "commentary": commentary,
        "manual_quotes": manual,
    }
    path = os.path.join(HERE, "rates.json")
    with open(path, "w") as f:
        json.dump(payload, f, separators=(",", ":"))
    log(f"wrote {path} ({os.path.getsize(path)/1e6:.2f} MB)")

    # A missing context series is loud but not fatal; a missing market is both.
    if context_errors and os.environ.get("GITHUB_ACTIONS"):
        detail = " · ".join(f"{e['series']}: {e['detail']}" for e in context_errors)
        print(f"::warning title=Context series unavailable::{detail}")

    missing = [m["code"] for m in MARKETS
               if m["code"] not in {x["code"] for x in out_markets}]
    if missing:
        log(f"FAILED markets: {', '.join(missing)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
